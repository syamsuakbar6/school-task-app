from io import BytesIO
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from sqlalchemy import select, delete, func
from sqlalchemy.orm import Session

from app.core.dependencies import DBSession, get_current_user
from app.core.security import hash_password
from app.models.class_model import Class, ClassMembership, TeacherClassAssignment
from app.models.submission import Submission
from app.models.task import Task
from app.models.user import User, UserRole
from app.schemas.user_schema import UserResponse
from pydantic import BaseModel, Field


router = APIRouter(prefix="/admin", tags=["Admin"])


# ── Guards ────────────────────────────────────────────────────────────────────

def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if str(current_user.role).lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Akses admin diperlukan.",
        )
    return current_user


# ── Schemas ───────────────────────────────────────────────────────────────────

class CreateStudentRequest(BaseModel):
    name: str = Field(min_length=3, max_length=100)
    nisn: str = Field(min_length=10, max_length=10, pattern=r'^\d{10}$')


class CreateTeacherRequest(BaseModel):
    name: str = Field(min_length=3, max_length=100)
    nip: str = Field(min_length=18, max_length=18, pattern=r'^\d{18}$')


class CreateClassRequest(BaseModel):
    name: str = Field(min_length=3, max_length=100)
    code: str = Field(min_length=1, max_length=20)


class UpdateClassRequest(BaseModel):
    name: str | None = Field(default=None, min_length=3, max_length=100)
    code: str | None = Field(default=None, min_length=1, max_length=20)


class StudentResponse(BaseModel):
    id: int
    name: str
    nisn: str | None
    role: str

    class Config:
        from_attributes = True


class TeacherResponse(BaseModel):
    id: int
    name: str
    nip: str | None
    role: str

    class Config:
        from_attributes = True


class ClassResponse(BaseModel):
    id: int
    name: str
    code: str | None
    teacher_id: int | None = None
    is_archived: bool = False
    archived_at: datetime | None = None

    class Config:
        from_attributes = True


class ClassWithStudentsResponse(BaseModel):
    id: int
    name: str
    code: str | None
    students: list[StudentResponse] = Field(default_factory=list)
    teachers: list[TeacherResponse] = Field(default_factory=list)
    task_count: int = 0
    submission_count: int = 0
    is_archived: bool = False
    archived_at: datetime | None = None

    class Config:
        from_attributes = True


class StudentImportResponse(BaseModel):
    total_rows: int
    assigned_count: int
    created_count: int
    skipped_count: int
    assigned: list[StudentResponse] = Field(default_factory=list)
    skipped: list[dict[str, str]] = Field(default_factory=list)


# ── User endpoints ────────────────────────────────────────────────────────────

@router.get("/users", response_model=list[UserResponse])
def list_all_users(
    db: DBSession,
    _: User = Depends(require_admin),
) -> list[UserResponse]:
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    return [UserResponse.model_validate(u) for u in users]


@router.post("/students", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_student(
    payload: CreateStudentRequest,
    db: DBSession,
    _: User = Depends(require_admin),
) -> UserResponse:
    # Cek duplikat NISN
    existing = db.scalar(select(User).where(User.nisn == payload.nisn))
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="NISN sudah terdaftar.",
        )

    # Password default = NISN
    student = User(
        name=payload.name.strip(),
        nisn=payload.nisn,
        role=UserRole.STUDENT.value,
        password=hash_password(payload.nisn),
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return UserResponse.model_validate(student)


@router.post("/teachers", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_teacher(
    payload: CreateTeacherRequest,
    db: DBSession,
    _: User = Depends(require_admin),
) -> UserResponse:
    existing = db.scalar(select(User).where(User.nip == payload.nip))
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="NIP sudah terdaftar.",
        )

    teacher = User(
        name=payload.name.strip(),
        nip=payload.nip,
        role=UserRole.TEACHER.value,
        password=hash_password(payload.nip),
    )
    db.add(teacher)
    db.commit()
    db.refresh(teacher)
    return UserResponse.model_validate(teacher)


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int,
    db: DBSession,
    current_admin: User = Depends(require_admin),
) -> None:
    if user_id == current_admin.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tidak bisa menghapus akun sendiri.",
        )

    user = db.scalar(select(User).where(User.id == user_id))
    if user is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User tidak ditemukan.",
        )

    # Hapus relasi kelas dulu sebelum hapus user
    db.execute(delete(ClassMembership).where(ClassMembership.student_id == user_id))
    db.execute(delete(TeacherClassAssignment).where(TeacherClassAssignment.teacher_id == user_id))
    db.delete(user)
    db.commit()


# ── Class endpoints ───────────────────────────────────────────────────────────

@router.get("/classes", response_model=list[ClassWithStudentsResponse])
def list_all_classes(
    db: DBSession,
    _: User = Depends(require_admin),
    include_archived: bool = False,
) -> list[ClassWithStudentsResponse]:
    statement = select(Class).order_by(Class.is_archived.asc(), Class.created_at.desc())
    if not include_archived:
        statement = statement.where(Class.is_archived.is_(False))
    classes = db.scalars(statement).all()
    result = []
    for c in classes:
        memberships = db.scalars(
            select(ClassMembership).where(ClassMembership.class_id == c.id)
        ).all()
        student_ids = [m.student_id for m in memberships]
        students = []
        if student_ids:
            students = db.scalars(
                select(User).where(User.id.in_(student_ids))
            ).all()
        assignments = db.scalars(
            select(TeacherClassAssignment).where(TeacherClassAssignment.class_id == c.id)
        ).all()
        teacher_ids = [a.teacher_id for a in assignments]
        teachers = []
        if teacher_ids:
            teachers = db.scalars(
                select(User).where(User.id.in_(teacher_ids))
            ).all()
        task_count = db.scalar(
            select(func.count()).select_from(Task).where(Task.class_id == c.id)
        ) or 0
        submission_count = db.scalar(
            select(func.count()).select_from(Submission).where(Submission.class_id == c.id)
        ) or 0
        result.append(ClassWithStudentsResponse(
            id=c.id,
            name=c.name,
            code=c.code,
            students=[StudentResponse.model_validate(s) for s in students],
            teachers=[TeacherResponse.model_validate(t) for t in teachers],
            task_count=task_count,
            submission_count=submission_count,
            is_archived=bool(c.is_archived),
            archived_at=c.archived_at,
        ))
    return result


@router.post("/classes", response_model=ClassResponse, status_code=status.HTTP_201_CREATED)
def create_class(
    payload: CreateClassRequest,
    db: DBSession,
    current_admin: User = Depends(require_admin),
) -> ClassResponse:
    normalized_name = _normalize_class_name(payload.name)
    normalized_code = payload.code.strip().upper()
    _assert_active_class_name_available(db, normalized_name=normalized_name)
    existing = db.scalar(select(Class).where(Class.code == normalized_code))
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Kode kelas sudah digunakan.",
        )

    new_class = Class(
        name=normalized_name,
        code=normalized_code,
        teacher_id=current_admin.id,
    )
    db.add(new_class)
    db.commit()
    db.refresh(new_class)
    return ClassResponse.model_validate(new_class)


@router.patch("/classes/{class_id}", response_model=ClassResponse)
def update_class(
    class_id: int,
    payload: UpdateClassRequest,
    db: DBSession,
    _: User = Depends(require_admin),
) -> ClassResponse:
    cls = _get_class_or_404(db, class_id=class_id)
    if cls.is_archived:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Kelas arsip harus dipulihkan sebelum diubah.",
        )

    if payload.name is not None:
        normalized_name = _normalize_class_name(payload.name)
        if normalized_name.lower() != cls.name.lower():
            _assert_active_class_name_available(
                db,
                normalized_name=normalized_name,
                exclude_class_id=class_id,
            )
        cls.name = normalized_name

    if payload.code is not None:
        normalized_code = payload.code.strip().upper()
        if normalized_code != (cls.code or "").upper():
            existing = db.scalar(
                select(Class).where(Class.code == normalized_code, Class.id != class_id)
            )
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Kode kelas sudah digunakan.",
                )
        cls.code = normalized_code

    db.add(cls)
    db.commit()
    db.refresh(cls)
    return ClassResponse.model_validate(cls)


@router.post("/classes/{class_id}/archive", response_model=ClassResponse)
def archive_class(
    class_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> ClassResponse:
    cls = _get_class_or_404(db, class_id=class_id)
    if not cls.is_archived:
        cls.is_archived = True
        cls.archived_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.add(cls)
        db.commit()
        db.refresh(cls)
    return ClassResponse.model_validate(cls)


@router.post("/classes/{class_id}/unarchive", response_model=ClassResponse)
def unarchive_class(
    class_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> ClassResponse:
    cls = _get_class_or_404(db, class_id=class_id)
    _assert_active_class_name_available(
        db,
        normalized_name=cls.name,
        exclude_class_id=class_id,
    )
    cls.is_archived = False
    cls.archived_at = None
    db.add(cls)
    db.commit()
    db.refresh(cls)
    return ClassResponse.model_validate(cls)


def _get_class_or_404(db: Session, *, class_id: int) -> Class:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Kelas tidak ditemukan.",
        )
    return cls


def _normalize_class_name(value: str) -> str:
    return " ".join(value.strip().split())


def _assert_active_class_name_available(
    db: Session,
    *,
    normalized_name: str,
    exclude_class_id: int | None = None,
) -> None:
    statement = select(Class).where(
        func.lower(Class.name) == normalized_name.lower(),
        Class.is_archived.is_(False),
    )
    if exclude_class_id is not None:
        statement = statement.where(Class.id != exclude_class_id)
    existing = db.scalar(statement)
    if existing is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nama kelas sudah digunakan.",
        )


@router.delete("/classes/{class_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_class(
    class_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> None:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Kelas tidak ditemukan.",
        )

    task_count = db.scalar(
        select(func.count()).select_from(Task).where(Task.class_id == class_id)
    ) or 0
    submission_count = db.scalar(
        select(func.count()).select_from(Submission).where(Submission.class_id == class_id)
    ) or 0
    if task_count or submission_count:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Kelas masih memiliki "
                f"{task_count} tugas dan {submission_count} pengumpulan. "
                "Hapus kelas hanya untuk kelas kosong."
            ),
        )

    db.execute(delete(ClassMembership).where(ClassMembership.class_id == class_id))
    db.execute(delete(TeacherClassAssignment).where(TeacherClassAssignment.class_id == class_id))
    db.delete(cls)
    db.commit()


# ── Assign endpoints ──────────────────────────────────────────────────────────

@router.get("/classes/{class_id}/students", response_model=list[StudentResponse])
def list_students_in_class(
    class_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> list[StudentResponse]:
    memberships = db.scalars(
        select(ClassMembership).where(ClassMembership.class_id == class_id)
    ).all()
    student_ids = [m.student_id for m in memberships]
    if not student_ids:
        return []
    students = db.scalars(select(User).where(User.id.in_(student_ids))).all()
    return [StudentResponse.model_validate(s) for s in students]


@router.post(
    "/classes/{class_id}/students/{student_id}",
    status_code=status.HTTP_201_CREATED,
)
def assign_student_to_class(
    class_id: int,
    student_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> dict:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan.")

    student = db.scalar(select(User).where(User.id == student_id))
    if student is None or str(student.role).lower() != UserRole.STUDENT.value:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan.")

    existing = db.scalar(
        select(ClassMembership).where(
            ClassMembership.student_id == student_id,
        )
    )
    if existing:
        if existing.class_id == class_id:
            detail = "Siswa sudah terdaftar di kelas ini."
        else:
            detail = "Siswa sudah terdaftar di kelas lain."
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=detail,
        )

    membership = ClassMembership(class_id=class_id, student_id=student_id)
    db.add(membership)
    db.commit()
    return {"message": "Siswa berhasil ditambahkan ke kelas."}


@router.delete(
    "/classes/{class_id}/students/{student_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def remove_student_from_class(
    class_id: int,
    student_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> None:
    membership = db.scalar(
        select(ClassMembership).where(
            ClassMembership.class_id == class_id,
            ClassMembership.student_id == student_id,
        )
    )
    if membership is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Siswa tidak terdaftar di kelas ini.",
        )
    db.delete(membership)
    db.commit()


@router.post(
    "/classes/{class_id}/students/import",
    response_model=StudentImportResponse,
)
async def import_students_to_class(
    class_id: int,
    file: UploadFile,
    db: DBSession,
    _: User = Depends(require_admin),
) -> StudentImportResponse:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan.")

    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File harus berformat .xlsx atau .xlsm.",
        )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            filename=BytesIO(await file.read()),
            read_only=True,
            data_only=True,
        )
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Dependency openpyxl belum terinstall di server.",
        ) from exc
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel tidak bisa dibaca.",
        ) from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return StudentImportResponse(
            total_rows=0,
            assigned_count=0,
            created_count=0,
            skipped_count=0,
        )

    header_map = {
        _normalize_excel_header(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    nisn_index = header_map.get("nisn")
    if nisn_index is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Kolom NISN tidak ditemukan.",
        )
    name_index = header_map.get("nama")
    if name_index is None:
        name_index = header_map.get("name")

    assigned: list[User] = []
    created_count = 0
    skipped: list[dict[str, str]] = []
    seen_nisn: set[str] = set()
    total_rows = 0

    for row_number, row in enumerate(rows, start=2):
        nisn = _cell_to_text(row[nisn_index] if nisn_index < len(row) else None)
        name = _cell_to_text(
            row[name_index] if name_index is not None and name_index < len(row) else None
        )
        if not nisn:
            continue

        total_rows += 1
        if nisn in seen_nisn:
            skipped.append({
                "row": str(row_number),
                "nisn": nisn,
                "reason": "NISN duplikat di file.",
            })
            continue
        seen_nisn.add(nisn)

        if len(nisn) != 10 or not nisn.isdigit():
            skipped.append({
                "row": str(row_number),
                "nisn": nisn,
                "reason": "NISN harus 10 digit angka.",
            })
            continue

        student = db.scalar(select(User).where(User.nisn == nisn))
        if student is None or str(student.role).lower() != UserRole.STUDENT.value:
            if len(name.strip()) < 3:
                skipped.append({
                    "row": str(row_number),
                    "nisn": nisn,
                    "reason": "Siswa tidak ditemukan dan kolom nama kosong.",
                })
                continue
            student = User(
                name=name.strip(),
                nisn=nisn,
                role=UserRole.STUDENT.value,
                password=hash_password(nisn),
            )
            db.add(student)
            db.flush()
            created_count += 1

        existing = db.scalar(
            select(ClassMembership).where(ClassMembership.student_id == student.id)
        )
        if existing:
            reason = (
                "Siswa sudah terdaftar di kelas ini."
                if existing.class_id == class_id
                else "Siswa sudah terdaftar di kelas lain."
            )
            skipped.append({
                "row": str(row_number),
                "nisn": nisn,
                "reason": reason,
            })
            continue

        db.add(ClassMembership(class_id=class_id, student_id=student.id))
        assigned.append(student)

    db.commit()
    return StudentImportResponse(
        total_rows=total_rows,
        assigned_count=len(assigned),
        created_count=created_count,
        skipped_count=len(skipped),
        assigned=[StudentResponse.model_validate(student) for student in assigned],
        skipped=skipped,
    )


def _normalize_excel_header(value: object) -> str:
    return "".join(str(value).strip().lower().split()).replace("_", "")


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value).strip().zfill(10)
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip().zfill(10)
    return str(value).strip()


@router.get("/classes/{class_id}/teachers", response_model=list[TeacherResponse])
def list_teachers_in_class(
    class_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> list[TeacherResponse]:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan.")

    assignments = db.scalars(
        select(TeacherClassAssignment).where(TeacherClassAssignment.class_id == class_id)
    ).all()
    teacher_ids = [a.teacher_id for a in assignments]
    if not teacher_ids:
        return []
    teachers = db.scalars(select(User).where(User.id.in_(teacher_ids))).all()
    return [TeacherResponse.model_validate(t) for t in teachers]


@router.post(
    "/classes/{class_id}/teachers/{teacher_id}",
    status_code=status.HTTP_201_CREATED,
)
def assign_teacher_to_class(
    class_id: int,
    teacher_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> dict:
    cls = db.scalar(select(Class).where(Class.id == class_id))
    if cls is None:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan.")

    teacher = db.scalar(select(User).where(User.id == teacher_id))
    if teacher is None or str(teacher.role).lower() != UserRole.TEACHER.value:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan.")

    existing = db.scalar(
        select(TeacherClassAssignment).where(
            TeacherClassAssignment.class_id == class_id,
            TeacherClassAssignment.teacher_id == teacher_id,
        )
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Guru sudah terdaftar di kelas ini.",
        )

    assignment = TeacherClassAssignment(class_id=class_id, teacher_id=teacher_id)
    db.add(assignment)
    db.commit()
    return {"message": "Guru berhasil ditambahkan ke kelas."}


@router.delete(
    "/classes/{class_id}/teachers/{teacher_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def remove_teacher_from_class(
    class_id: int,
    teacher_id: int,
    db: DBSession,
    _: User = Depends(require_admin),
) -> None:
    assignment = db.scalar(
        select(TeacherClassAssignment).where(
            TeacherClassAssignment.class_id == class_id,
            TeacherClassAssignment.teacher_id == teacher_id,
        )
    )
    if assignment is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Guru tidak terdaftar di kelas ini.",
        )
    db.delete(assignment)
    db.commit()
